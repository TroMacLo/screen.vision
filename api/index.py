from typing import Any, List
from pydantic import BaseModel
from dotenv import load_dotenv
from fastapi import FastAPI, File, Request as FastAPIRequest, UploadFile
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from openai import OpenAI
from io import BytesIO
import base64
import csv
import os

from google import genai
from google.genai import types

from .utils.stream import stream_text
from .utils.gemini import convert_openai_to_gemini, stream_gemini

# Monkeypatch ThinkingConfig to allow extra fields like thinking_level
types.ThinkingConfig.model_config["extra"] = "allow"
types.ThinkingConfig.model_rebuild(force=True)

load_dotenv(".env.local")

limiter = Limiter(key_func=get_remote_address)
app = FastAPI()
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

is_production = (
    os.getenv("RAILWAY_ENVIRONMENT_NAME") == "production"
    or os.getenv("VERCEL_ENV") == "production"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://screen.vision", "https://www.screen.vision"]
    if is_production
    else ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class MessagesRequest(BaseModel):
    messages: List[Any]


class FileContextItem(BaseModel):
    name: str
    size: int
    mime_type: str
    analysis: str


class FileContextResponse(BaseModel):
    files: List[FileContextItem]


def _truncate_text(text: str, max_length: int = 12000) -> str:
    trimmed = text.strip()
    if len(trimmed) <= max_length:
        return trimmed
    return f"{trimmed[:max_length]}\n...[truncated]"


def _analyze_text_file(contents: bytes) -> str:
    decoded = contents.decode("utf-8", errors="ignore")
    return _truncate_text(decoded)


def _analyze_csv_file(contents: bytes) -> str:
    decoded = contents.decode("utf-8", errors="ignore")
    reader = csv.reader(decoded.splitlines())
    rows = []
    for i, row in enumerate(reader):
        if i >= 120:
            rows.append("...[truncated]")
            break
        rows.append(", ".join(cell.strip() for cell in row))
    return _truncate_text("\n".join(rows))


def _analyze_pdf_file(contents: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(contents))
    pages = []
    for index, page in enumerate(reader.pages[:25]):
        page_text = page.extract_text() or ""
        pages.append(f"Page {index + 1}:\n{page_text}")
    return _truncate_text("\n\n".join(pages))


def _analyze_docx_file(contents: bytes) -> str:
    from docx import Document

    doc = Document(BytesIO(contents))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return _truncate_text("\n".join(paragraphs))


def _analyze_spreadsheet_file(contents: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(contents), data_only=True, read_only=True)
    snippets = []

    try:
        for sheet in workbook.worksheets[:5]:
            snippets.append(f"Sheet: {sheet.title}")
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                if row_count >= 120:
                    snippets.append("...[truncated]")
                    break
                values = [str(cell) if cell is not None else "" for cell in row]
                snippets.append(" | ".join(values))
                row_count += 1

        return _truncate_text("\n".join(snippets))
    finally:
        workbook.close()


def _analyze_image_file(contents: bytes, mime_type: str, filename: str) -> str:
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return f"Image attached: {filename}. Set OPENAI_API_KEY for automated visual analysis."

    client = OpenAI(api_key=api_key)
    data_url = f"data:{mime_type};base64,{base64.b64encode(contents).decode('utf-8')}"

    result = client.chat.completions.create(
        model="gpt-5-mini-2025-08-07",
        messages=[
            {
                "role": "system",
                "content": "Summarize this uploaded image for task guidance. Focus on actionable, concise details.",
            },
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": f"Analyze image: {filename}"},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            },
        ],
        reasoning_effort="minimal",
    )

    return _truncate_text(result.choices[0].message.content or "")


def _analyze_uploaded_file(file: UploadFile, contents: bytes) -> str:
    mime_type = file.content_type or "application/octet-stream"
    filename = (file.filename or "uploaded-file").lower()

    is_text = mime_type.startswith("text/") or filename.endswith(
        (
            ".md",
            ".txt",
            ".json",
            ".yaml",
            ".yml",
            ".xml",
            ".log",
            ".html",
            ".js",
            ".ts",
            ".tsx",
            ".py",
            ".sql",
        )
    )

    if mime_type.startswith("image/"):
        return _analyze_image_file(contents, mime_type, file.filename or "image")

    if mime_type == "application/pdf" or filename.endswith(".pdf"):
        return _analyze_pdf_file(contents)

    if filename.endswith(".docx"):
        return _analyze_docx_file(contents)

    if filename.endswith((".xlsx", ".xlsm", ".xltx")):
        return _analyze_spreadsheet_file(contents)

    if mime_type in ["text/csv", "application/csv"] or filename.endswith(".csv"):
        return _analyze_csv_file(contents)

    if is_text:
        return _analyze_text_file(contents)

    return (
        f"Attached file {file.filename or 'uploaded-file'} ({mime_type}). "
        "This file type is not fully parseable yet, but the assistant should still consider that it was provided."
    )


@app.post("/api/file-context", response_model=FileContextResponse)
@limiter.limit("20/minute;250/hour")
async def analyze_file_context(
    request: FastAPIRequest,
    files: List[UploadFile] = File(...),
):
    analyzed_files: List[FileContextItem] = []

    for file in files:
        contents = await file.read()
        if len(contents) > 30 * 1024 * 1024:
            raise ValueError(f"File {file.filename} exceeds the 30MB limit")

        try:
            analysis = _analyze_uploaded_file(file, contents)
        except Exception as exc:  # pragma: no cover - resilient per-file fallback
            analysis = (
                f"Attached file {file.filename or 'uploaded-file'} could not be fully analyzed. "
                f"Error: {str(exc)}"
            )

        analyzed_files.append(
            FileContextItem(
                name=file.filename or "uploaded-file",
                size=len(contents),
                mime_type=file.content_type or "application/octet-stream",
                analysis=analysis,
            )
        )

    return FileContextResponse(files=analyzed_files)


@app.post("/api/step")
@limiter.limit("20/minute;300/hour")
async def handle_step_chat(request: FastAPIRequest, body: MessagesRequest):
    client = OpenAI()

    stream = client.chat.completions.create(
        messages=body.messages,
        model="gpt-5-mini-2025-08-07",
        stream=True,
        reasoning_effort="low",
    )

    response = StreamingResponse(
        stream_text(stream, {}),
        media_type="text/event-stream",
    )

    return response


@app.post("/api/help")
@limiter.limit("8/minute;100/hour")
async def handle_help_chat(request: FastAPIRequest, body: MessagesRequest):
    client = OpenAI()

    stream = client.chat.completions.create(
        messages=body.messages,
        model="gpt-5-mini-2025-08-07",
        stream=True,
        reasoning_effort="low",
    )

    response = StreamingResponse(
        stream_text(stream, {}),
        media_type="text/event-stream",
    )

    return response


@app.post("/api/check")
@limiter.limit("30/minute;500/hour")
async def handle_check_chat(request: FastAPIRequest, body: MessagesRequest):
    gemini_api_key = os.environ.get("GEMINI_API_KEY")

    if gemini_api_key:
        client = genai.Client(
            vertexai=True,
            api_key=gemini_api_key,
        )
        model = "gemini-3-flash-preview"

        system_instruction_parts = []
        for msg in body.messages:
            if msg.get("role") == "system":
                content = msg.get("content")
                if isinstance(content, str):
                    system_instruction_parts.append(types.Part.from_text(text=content))
                elif isinstance(content, list):
                    for part in content:
                        if part.get("type") == "text":
                            system_instruction_parts.append(
                                types.Part.from_text(text=part.get("text"))
                            )

        system_instruction = (
            types.Content(parts=system_instruction_parts)
            if system_instruction_parts
            else None
        )

        contents = convert_openai_to_gemini(body.messages)

        generate_content_config = types.GenerateContentConfig(
            system_instruction=system_instruction,
            thinking_config=types.ThinkingConfig(
                thinking_level="MINIMAL",
            ),
        )

        stream = client.models.generate_content_stream(
            model=model,
            contents=contents,
            config=generate_content_config,
        )

        response = StreamingResponse(
            stream_gemini(stream),
            media_type="text/event-stream",
        )
        return response
    else:
        client = OpenAI(
            base_url="https://openrouter.ai/api/v1",
            api_key=os.environ.get("OPENROUTER_API_KEY"),
        )
        kwargs = {
            "messages": body.messages,
            "model": "google/gemini-3-flash-preview",
            "extra_body": {
                "provider": {
                    "order": ["Google AI Studio"],
                    "allow_fallbacks": True,
                }
            },
            "reasoning_effort": "minimal",
            "stream": True,
        }

    stream = client.chat.completions.create(**kwargs)

    response = StreamingResponse(
        stream_text(stream, {}),
        media_type="text/event-stream",
    )

    return response


@app.post("/api/coordinates")
@limiter.limit("15/minute;200/hour")
async def handle_coordinate_chat(request: FastAPIRequest, body: MessagesRequest):
    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=os.environ.get("OPENROUTER_API_KEY"),
    )

    stream = client.chat.completions.create(
        messages=body.messages,
        model="qwen/qwen3-vl-30b-a3b-instruct",
        extra_body={"provider": {"order": ["Fireworks"], "allow_fallbacks": True}},
        stream=True,
    )

    response = StreamingResponse(
        stream_text(stream, {}),
        media_type="text/event-stream",
    )

    return response
